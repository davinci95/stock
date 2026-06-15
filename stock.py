"""
LỌC CỔ PHIẾU PHI THƯỜNG – HOSE
pip install vnstock openpyxl
python loc_co_phieu.py
"""
import time
from datetime import datetime
from vnstock import Fundamental
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TICKERS = [
    'GEG','VID','PNJ','CVT','SAM','KHG','FIT','HHS','LBM','CNG','SC5','ELC',
    'LSS','VRC','FRT','PMG','SAV','HHP','SSB','PGI','PET','HVN','KLB','TDP',
    'PLP','HTG','CRC','TDM','SCR','DHM','EVE','TLD','SVD','BWE','TRC','IJC',
    'YBM','TNT','DBT','PPC','SJD','ABR','SSC','CLW','TDG','ACB','TMS','KOS',
    'MSH','CHP','HPA','ADP','TCX','DAH','PTC','ACL','TRA','MSB','VPG','BKG',
    'CIG','SBG','PTL','DGW','CMG','TSC','SIP','SHA','HPX','SSI','HRC','BIC',
    'SPM','PIT','SMC','CKG','MHC','TYA','DSN','ICT','BMC','TAL','STB','HTN',
    'DMC','TMT','KBC','PDV','CTF','DRH','VCG','QNP','HTI','LIX','TDW','TVB',
    'DHG','BID','DCM','IDI','VHC','CSV','EVF','NAB','SAB','TN1','BTP','AAT',
    'SGT','MCH','TNH','BCE','VCA','ABT','CTD','PVD','ASP','TCH','DHA','PC1',
    'L10','HAS','PTB','VCK','DVP','JVC','NKG','AAA','HAR','BMP','TCI','BSR',
    'CLC','VPL','PVT','ASG','KHP','DIG','TBC','SGR','DRC','SJS','VCF','ASM',
    'VIC','SBV','HUB','VCI','OGC','C47','NTC','KSB','VSI','VJC','VPI','LGC',
    'HT1','C32','FTS','VDP','VPX','TLG','TEG','DXV','TCM','BSI','PGC','ITC',
    'TCB','VND','D2D','AAM','ITD','CTI','NHT','TPB','LPB','TCD','SHI','PJT',
    'VFG','VRE','VNG','TV2','VPS','MWG','NNC','NHH','TMP','CSM','DAT','TIX',
    'VPB','STG','SBA','ANV','ADG','VSH','NAF','SMA','FDC','TSA','ADS','SCS',
    'S4A','VCB','QCG','VHM','CTR','TDC','HCD','KMR','DXS','AST','HU1','ACC',
    'HVH','GEE','CRE','HNA','VIX','SHP','VNM','SBT','HTV','HMC','SZL','GAS',
    'SRF','TCR','BTT','YEG','VAB','MBB','VTP','SZC','VTB','HHV','NHA','TNI',
    'DPR','GVR','APH','TCO','HAH','LCG','VPH','NVT','VTO','CMV','PVP','HQC',
    'BCM','NSC','DXG','REE','DBC','AGR','DTL','DLG','SKG','ORS','TCT','SVT',
    'GHC','NCT','FIR','VIB','NLG','APG','SHB','CMX','RAL','SFG','TVS','TVT',
    'CDC','SFC','HDB','DBD','VGC','RYG','DC4','COM','CCL','FCN','HTL','GDT',
    'IMP','MSN','CII','OCB','TCL','DTT','CTG','UIC','DPM','NTL','BRC','AFX',
    'DCL','ACG','MDG','POW','CLL','LGL','BFC','GIL','PHR','FPT','DGC','NVL',
    'LAF','TPC','LHG','VSC','MCP','VNS','THG','BMI','SFI','FCM','KDC','ABS',
    'DPG','SGN','GMH','PGV','SRC','NT2','DSC','TNC','PAC','GTA','CRV','ANT',
    'DHC','VOS','TLH','CCC','DRL','STK','PNC','TDH','EIB','NO1','TIP','GEX',
    'MIG','FMC','GEL','HAX','TTF','MCM','PDR','TTA','ST8','KDH','PAN','SVC',
    'EVG','HSL','NBB','VMD','SMB','AGG','VPD','DTA','BVH','VVS','HAG','VNL',
    'VIP','HDG','PGD','PDN','HCM','GSP','LM8','OPC','HID','TTE','BCG','PLX',
    'HPG','GMD','VDS','NAV','HAP','CCI','VNE','DQC','PHC','HII','LDG','CTS',
    'BHN','ILB','HDC','DSE','BAF','HSG',
]

DELAY = 0.4
THIN  = Side(style='thin', color='CCCCCC')
BDR   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
REC_BG = {
    "🌟 MUA MANH": "1A6B2A", "✅ MUA": "2E8B57",
    "🔍 THEO DOI": "B8860B", "⚠️ TRUNG TINH": "CD853F", "❌ TRANH": "8B0000",
}

# Lọc cứng – KHÔNG dùng profit_growth và net_margin
HARD_FILTER = {
    'roe_min':           15,
    'de_max':           150,
    'pe_max':            35,
    'revenue_growth_min': 0,
}

def get_val(df, item_id):
    time_cols = [c for c in df.columns if c not in ('item','item_id')]
    if not time_cols: return None
    row = df[df['item_id'] == item_id]
    if row.empty: return None
    try:
        v = float(row.iloc[0][time_cols[0]])
        return None if v != v else v
    except: return None

def fetch(ticker):
    try:
        df = Fundamental().equity(ticker).ratio(period='quarter')
        if df is None or df.empty: return {}
        pe = get_val(df, 'p_e')
        rg = get_val(df, 'net_revenue')
        peg = round(pe / rg, 2) if pe and rg and rg > 0 else None
        return {
            'pe':           pe,
            'pb':           get_val(df, 'p_b'),
            'ps':           get_val(df, 'p_s'),
            'ev_ebitda':    get_val(df, 'ev_ebitda'),
            'peg':          peg,
            'eps':          get_val(df, 'trailing_eps'),
            'roe':          get_val(df, 'roe_trailling'),
            'roa':          get_val(df, 'roa_trailling'),
            'gross_margin': get_val(df, 'gross_profit_margin'),
            'ebit_margin':  get_val(df, 'ebit_margin'),
            'revenue_growth': rg,
            'gross_profit_growth': get_val(df, 'gross_profit'),
            'de_ratio':     get_val(df, 'debt_to_equity'),
            'current_ratio':get_val(df, 'short_term_ratio'),
            'interest_cov': get_val(df, 'interest_coverage'),
            'ocf_per_share':get_val(df, 'cash_flow_per_share_cps'),
            'div_yield':    get_val(df, 'dividend_yield'),
            'beta':         get_val(df, 'beta'),
            'bvps':         get_val(df, 'book_value_per_share_bvps'),
        }
    except: return {}

def passes_hard_filter(d):
    if not d: return False, "Khong co du lieu"
    checks = [
        (d.get('roe'),            lambda v: v >= HARD_FILTER['roe_min'],            f"ROE < {HARD_FILTER['roe_min']}%"),
        (d.get('de_ratio'),       lambda v: v <= HARD_FILTER['de_max'],             f"D/E > {HARD_FILTER['de_max']}%"),
        (d.get('pe'),             lambda v: 0 < v <= HARD_FILTER['pe_max'],         f"P/E > {HARD_FILTER['pe_max']} hoac am"),
        (d.get('revenue_growth'), lambda v: v >= HARD_FILTER['revenue_growth_min'], f"DT tang truong am"),
    ]
    for val, cond, reason in checks:
        if val is None: return False, f"Thieu du lieu ({reason})"
        if not cond(val): return False, reason
    return True, "Dat tieu chi"

def score_stock(d):
    s = 0; notes = []

    # Tăng trưởng DT (20đ)
    rg = d.get('revenue_growth')
    if rg is not None:
        if   rg >= 30: s+=20; notes.append(f"✅ DT tang {rg:.1f}% (xuat sac)")
        elif rg >= 15: s+=14; notes.append(f"✅ DT tang {rg:.1f}%")
        elif rg >= 0:  s+=6;  notes.append(f"⚠️ DT tang {rg:.1f}% (thap)")
    else: notes.append("❓ Khong co tang truong DT")

    # ROE trailing (20đ)
    roe = d.get('roe')
    if roe is not None:
        if   roe >= 25: s+=20; notes.append(f"✅ ROE {roe:.1f}% (xuat sac)")
        elif roe >= 20: s+=15; notes.append(f"✅ ROE {roe:.1f}% (tot)")
        elif roe >= 15: s+=10; notes.append(f"✅ ROE {roe:.1f}%")
        else:                  notes.append(f"⚠️ ROE {roe:.1f}% (<15%)")
    else: notes.append("❓ Khong co ROE")

    # Biên gộp – proxy Moat (15đ)
    gm = d.get('gross_margin')
    if gm is not None:
        if   gm >= 40: s+=15; notes.append(f"✅ Bien gop {gm:.1f}% (loi the manh)")
        elif gm >= 25: s+=10; notes.append(f"✅ Bien gop {gm:.1f}%")
        elif gm >= 15: s+=5;  notes.append(f"⚠️ Bien gop {gm:.1f}% (trung binh)")
        else:                  notes.append(f"❌ Bien gop {gm:.1f}% (thap)")
    else: notes.append("❓ Khong co bien gop")

    # OCF dương (15đ)
    ocf = d.get('ocf_per_share')
    if ocf is not None:
        if ocf > 0: s+=15; notes.append(f"✅ OCF/cp duong ({ocf:,.0f})")
        else:               notes.append(f"❌ OCF/cp am ({ocf:,.0f})")
    else: notes.append("❓ Khong co OCF")

    # D/E (10đ)
    de = d.get('de_ratio')
    if de is not None:
        if   de < 30:  s+=10; notes.append(f"✅ D/E {de:.0f}% (rat an toan)")
        elif de < 80:  s+=7;  notes.append(f"✅ D/E {de:.0f}%")
        elif de < 150: s+=4;  notes.append(f"⚠️ D/E {de:.0f}% (chap nhan)")
        else:                  notes.append(f"❌ D/E {de:.0f}% (cao)")
    else: notes.append("❓ Khong co D/E")

    # P/E (10đ)
    pe = d.get('pe')
    if pe is not None and pe > 0:
        if   pe < 10:  s+=10; notes.append(f"✅ P/E {pe:.1f} (rat re)")
        elif pe < 15:  s+=8;  notes.append(f"✅ P/E {pe:.1f} (re)")
        elif pe < 20:  s+=6;  notes.append(f"✅ P/E {pe:.1f} (hop ly)")
        elif pe < 35:  s+=3;  notes.append(f"⚠️ P/E {pe:.1f} (cao)")
    else: notes.append("❓ Khong co P/E")

    # PEG (10đ)
    peg = d.get('peg')
    if peg is not None and peg > 0:
        if   peg < 0.5: s+=10; notes.append(f"✅ PEG {peg:.2f} (rat hap dan)")
        elif peg < 1.0: s+=7;  notes.append(f"✅ PEG {peg:.2f} (<1)")
        elif peg < 1.5: s+=4;  notes.append(f"⚠️ PEG {peg:.2f}")
        else:                   notes.append(f"❌ PEG {peg:.2f} (dat)")
    else: notes.append("❓ Khong tinh duoc PEG")

    return s, notes

def recommend(s):
    if s >= 80: return "🌟 MUA MANH"
    if s >= 65: return "✅ MUA"
    if s >= 50: return "🔍 THEO DOI"
    if s >= 35: return "⚠️ TRUNG TINH"
    return "❌ TRANH"

def wh(ws, r, c, v, bg="1F3864", fg="FFFFFF", sz=9, bold=True, wrap=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font      = Font(name="Arial", bold=bold, color=fg, size=sz)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    cell.border    = BDR

def wd(ws, r, c, v, fmt=None, bold=False, bg=None, fg="000000", left=False):
    cell = ws.cell(row=r, column=c, value=v if v is not None else "N/A")
    cell.font      = Font(name="Arial", size=9, bold=bold, color=fg)
    cell.alignment = Alignment(horizontal="left" if left else "center", vertical="center")
    cell.border    = BDR
    if bg:  cell.fill = PatternFill("solid", fgColor=bg)
    if fmt and v is not None: cell.number_format = fmt

def build_excel(all_rows, pass_rows, path):
    wb = Workbook()

    # ═══ Sheet 1: Mã đạt tiêu chí ════════════════════════════
    ws1 = wb.active
    ws1.title = "DAT TIEU CHI"
    ws1.merge_cells("A1:M1")
    c = ws1["A1"]
    c.value = f"CO PHIEU DAT TIEU CHI PHI THUONG – HOSE  |  {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    c.font  = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    c.fill  = PatternFill("solid", fgColor="0D1F3C")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 26

    H1 = [
        ("Ma CP",10),("Diem",8),("Khuyen nghi",14),
        ("P/E",8),("PEG",8),("P/B",8),
        ("ROE\ntrail(%)",10),("Bien\ngop(%)",10),
        ("DT\ntang(%)",10),("D/E(%)",9),
        ("EPS",12),("OCF/cp",10),("Beta",7),
    ]
    for ci,(h,w) in enumerate(H1,1):
        wh(ws1,2,ci,h,wrap=True)
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.row_dimensions[2].height = 34
    ws1.freeze_panes = "A3"

    for i,row in enumerate(pass_rows,3):
        d=row['data']; s=row['score']; rec=row['rec']
        alt = "F0FFF4" if i%2==0 else "FFFFFF"
        vals=[row['ticker'],s,rec,
              d.get('pe'),d.get('peg'),d.get('pb'),
              d.get('roe'),d.get('gross_margin'),
              d.get('revenue_growth'),d.get('de_ratio'),
              d.get('eps'),d.get('ocf_per_share'),d.get('beta')]
        fmts=[None,None,None,"0.0","0.00","0.0","0.0","0.0","0.0","0.0","#,##0","#,##0","0.00"]
        for ci,(val,fmt) in enumerate(zip(vals,fmts),1):
            bg=alt; fg="000000"
            if ci==2:
                bg=("1A6B2A" if s>=80 else "2E8B57" if s>=65 else "B8860B"); fg="FFFFFF"
            elif ci==3:
                bg=REC_BG.get(rec,alt); fg="FFFFFF"
            wd(ws1,i,ci,val,fmt=fmt,bold=(ci==1),bg=bg,fg=fg,left=(ci==1))
        ws1.row_dimensions[i].height = 15

    # ═══ Sheet 2: Toàn bộ kết quả ════════════════════════════
    ws2 = wb.create_sheet("TOAN BO")
    ws2.merge_cells("A1:N1")
    c = ws2["A1"]
    c.value = f"TOAN BO {len(all_rows)} MA HOSE – {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    c.font  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    c.fill  = PatternFill("solid", fgColor="2E4057")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 22

    H2 = [
        ("Ma CP",8),("Loc cung",10),("Ly do",24),("Diem",7),("Khuyen nghi",13),
        ("P/E",7),("PEG",7),("P/B",7),("ROE(%)",8),
        ("Bien gop(%)",10),("DT tang(%)",10),
        ("D/E(%)",8),("EPS",11),("Beta",7),
    ]
    for ci,(h,w) in enumerate(H2,1):
        wh(ws2,2,ci,h,wrap=True)
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[2].height = 34
    ws2.freeze_panes = "A3"

    for i,row in enumerate(all_rows,3):
        d=row['data']; s=row['score']; rec=row['rec']
        passed=row['passed']; reason=row['reason']
        alt = "F2F7FC" if i%2==0 else "FFFFFF"
        vals=[row['ticker'],
              "DAT" if passed else "LOAI", reason,
              s if passed else "", rec if passed else "",
              d.get('pe'),d.get('peg'),d.get('pb'),
              d.get('roe'),d.get('gross_margin'),d.get('revenue_growth'),
              d.get('de_ratio'),d.get('eps'),d.get('beta')]
        fmts=[None,None,None,None,None,"0.0","0.00","0.0","0.0","0.0","0.0","0.0","#,##0","0.00"]
        for ci,(val,fmt) in enumerate(zip(vals,fmts),1):
            bg=alt; fg="000000"
            if ci==2:
                bg = "D5F5E3" if passed else "FDECEA"
            elif ci==4 and passed:
                bg=("1A6B2A" if s>=80 else "2E8B57" if s>=65 else "B8860B"); fg="FFFFFF"
            elif ci==5 and passed:
                bg=REC_BG.get(rec,alt); fg="FFFFFF"
            wd(ws2,i,ci,val,fmt=fmt,bold=(ci==1),bg=bg,fg=fg,left=(ci<=3))
        ws2.row_dimensions[i].height = 15

    # ═══ Sheet 3: Chi tiết điểm ══════════════════════════════
    ws3 = wb.create_sheet("CHI TIET DIEM")
    ws3.merge_cells("A1:D1")
    c = ws3["A1"]
    c.value = "CHI TIET CHAM DIEM – MA DAT TIEU CHI"
    c.font  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    c.fill  = PatternFill("solid", fgColor="0D1F3C")
    c.alignment = Alignment(horizontal="center")
    ws3.row_dimensions[1].height = 22
    for ci,h in enumerate(["Ma CP","Diem","Khuyen nghi","Phan tich chi tiet"],1):
        wh(ws3,2,ci,h)
    ws3.column_dimensions["A"].width = 8
    ws3.column_dimensions["B"].width = 8
    ws3.column_dimensions["C"].width = 14
    ws3.column_dimensions["D"].width = 100
    ws3.row_dimensions[2].height = 16

    for i,row in enumerate(pass_rows,3):
        s=row['score']; rec=row['rec']
        alt="F0FFF4" if i%2==0 else "FFFFFF"
        score_bg="1A6B2A" if s>=80 else "2E8B57" if s>=65 else "B8860B"
        wd(ws3,i,1,row['ticker'],bold=True,bg=alt)
        wd(ws3,i,2,s,bg=score_bg,fg="FFFFFF")
        wd(ws3,i,3,rec,bg=REC_BG.get(rec,alt),fg="FFFFFF")
        c=ws3.cell(row=i,column=4,value="  |  ".join(row['notes']))
        c.font=Font(name="Arial",size=8); c.border=BDR
        c.fill=PatternFill("solid",fgColor=alt)
        c.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
        ws3.row_dimensions[i].height = 22

    wb.save(path)

def main():
    all_rows = []; pass_rows = []

    for ticker in TICKERS:
        d = fetch(ticker)
        passed, reason = passes_hard_filter(d)
        if passed:
            s, notes = score_stock(d)
            rec = recommend(s)
            row = dict(ticker=ticker,data=d,score=s,notes=notes,rec=rec,passed=True,reason=reason)
            all_rows.append(row); pass_rows.append(row)
        else:
            all_rows.append(dict(ticker=ticker,data=d,score=0,notes=[],rec="",passed=False,reason=reason))
        time.sleep(DELAY)

    pass_rows.sort(key=lambda x: x['score'], reverse=True)
    all_rows.sort(key=lambda x: (-x['score'], x['ticker']))

    out = f"loc_co_phieu_hose_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    build_excel(all_rows, pass_rows, out)

if __name__ == "__main__":
    main()